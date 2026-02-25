import openpyxl
from openpyxl.utils import column_index_from_string

# Exact sheet names — em-dashes, not hyphens
SHEET_MAP = {
    'Phase 1 — Taxonomy Audit': {
        'data_key': 'tags',
        'row_source': 'dict_values',
        'columns': {
            'A': 'tag_id',        'B': 'tag_title',
            'C': 'parent_tag',    'D': 'depth_level',
            'F': 'asset_count',   'G': 'status',
            'H': 'cloud_notes',   'I': 'recommended_map',
            'K': 'l1_id',         'L': 'l1_desc',
            'M': 'l1_title',      'N': 'l2_id',
            'O': 'l2_desc',       'P': 'l2_title',
            'Q': 'l3_id',         'R': 'l3_desc',
            'S': 'l3_title',      'T': 'l4_id',
            'U': 'l4_title',      'V': 'l4_desc',
            'W': 'full_tag_path', 'X': 'tag_label',
        }
    },
    'Phase 2 — Metadata Schema': {
        'data_key': 'metadata_fields',
        'row_source': 'dict_values',
        'columns': {
            'A': 'field_name',
            'C': 'data_type',
            'F': 'namespace',
            'H': 'current_usage_count',
        }
    },
    'Phase 3 — Workflow Extraction': {
        'data_key': 'workflows',
        'row_source': 'list',
        'columns': {
            'A': 'step_number', 'B': 'step_name',
            'C': 'step_type',   'E': 'fields_affected',
            'F': 'tags_used',   'G': 'conditions',
        }
    },
    'Phase 4 — Folder Redesign': {
        'data_key': 'folders',
        'row_source': 'dict_values',
        'columns': {
            'A': 'folder_path',    'B': 'folder_name',
            'C': 'depth_level',    'D': 'parent_folder',
            'E': 'child_count',    'F': 'asset_count',
            'G': 'is_metadata_like',
        }
    },
    'Phase 5 — Namespace Validation': {
        'data_key': 'namespaces',
        'row_source': 'dict_values',
        'columns': {
            'A': 'uri',              'B': 'prefix',
            'C': 'namespace_id',     'E': 'namespace_type',
            'F': 'used_in',          'G': 'cloud_support',
            'H': 'fields_in_namespace',
            'I': 'migration_strategy',
            'J': 'effort',           'K': 'timeline_days',
        }
    },
}


def write_all_phases(harvest: dict, workbook_path: str):
    """
    Write all phase data from harvest dict into the workbook.
    Reads SHEET_MAP to know which sheet, which column, which key.
    Starts writing at row 4. Never touches AI BOT or MANUAL columns.
    Saves back to workbook_path when done.
    """
    # Clear stale rows from all phase sheets before writing
    clear_phase_data(workbook_path)

    print(f"   [>>] Loading workbook: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path)

    for sheet_name, config in SHEET_MAP.items():

        if sheet_name not in wb.sheetnames:
            print(f"   [!] Sheet not found, skipping: {sheet_name}")
            continue

        ws = wb[sheet_name]
        data_key    = config['data_key']
        row_source  = config['row_source']
        col_map     = config['columns']

        # Get the data from harvest — handle missing keys gracefully
        raw_data = harvest.get(data_key)
        if not raw_data:
            print(f"   [!] No data for {sheet_name} "
                  f"(harvest['{data_key}'] is empty)")
            continue

        # Normalize to a list of dicts regardless of source type
        if row_source == 'dict_values':
            rows = list(raw_data.values())
        elif row_source == 'list':
            rows = raw_data
        else:
            rows = list(raw_data)

        # Write rows starting at row 4
        write_count = 0
        for i, row_dict in enumerate(rows):
            excel_row = 4 + i
            for col_letter, harvest_key in col_map.items():
                col_idx = column_index_from_string(col_letter)
                value   = row_dict.get(harvest_key, '')
                # Write None as empty string — keeps cells clean
                ws.cell(row=excel_row, column=col_idx,
                        value=value if value is not None else '')
            write_count += 1

        print(f"   [ok] {sheet_name}: {write_count} rows written")

    wb.save(workbook_path)
    print(f"   [saved] Workbook saved: {workbook_path}")


def clear_phase_data(workbook_path: str, phase: str = 'all'):
    """
    Clear data rows (row 4 onward) from phase sheets before a re-run.
    Preserves rows 1-3 (title, headers, source labels).
    Useful when re-running JCRUNCH against a new package.
    """
    wb = openpyxl.load_workbook(workbook_path)

    sheets_to_clear = []
    if phase == 'all':
        sheets_to_clear = list(SHEET_MAP.keys())
    else:
        phase_map = {
            '1': 'Phase 1 — Taxonomy Audit',
            '2': 'Phase 2 — Metadata Schema',
            '3': 'Phase 3 — Workflow Extraction',
            '4': 'Phase 4 — Folder Redesign',
            '5': 'Phase 5 — Namespace Validation',
        }
        if phase in phase_map:
            sheets_to_clear = [phase_map[phase]]

    for sheet_name in sheets_to_clear:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row >= 4:
            ws.delete_rows(4, ws.max_row - 3)
        print(f"   [ok] Cleared: {sheet_name}")

    wb.save(workbook_path)
